import base64
import copy
import csv
import hmac
import json
import subprocess
import time
import requests
import openpyxl
import os
import pandas as pd
import random
import string
from pywifi import ControlConnection
from dongle_lte_api import Dongle
from app.config import HOME_PACKAGE, HOME_TMP, get_logger, USER_DATA_DIR, ALL_USER_DATA_DIR, NETWORK_PASSWORD, \
    LIST_NETWORK, PATH_OF_AIRPORT

logger = get_logger(__name__)


def read_xlsx_file(dir_file: str, column_mapping: dict = None, sheet_name: str = None) -> list:
    wb = openpyxl.load_workbook(dir_file)
    if not sheet_name:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]  # the 1st sheet at index 0

    max_column = ws.max_column
    max_row = ws.max_row

    raw_headers = [ws.cell(row=1, column=ci).value for ci in range(1, max_column + 1)]  # ci aka column_index
    raw_headers = list(filter(None, raw_headers))  # remove None column out of header list

    v_fields = [h and column_mapping.get(h.strip()) for h in
                raw_headers]  # h aka header, ensure header is not null to strip and no error is thrown
    raw_v_rows = []  # raw_v_rows aka raw vehicle rows
    col_count = len(raw_headers)
    for ri in range(2, max_row + 1):  # ri aka row_index - we skip the 1st row which is the header rows
        values = [ws.cell(row=ri, column=ci).value for ci in range(1, col_count + 1)]  # ci aka column_index
        rvr = dict(zip(v_fields, values))  # rvr aka raw_vehicle_row
        raw_v_rows.append(rvr)
    return raw_v_rows


def read_csv_file(dir_file: str, column_mapping: dict = None) -> list:
    raw_v_rows = []  # raw_v_rows aka raw vehicle rows
    # region read csv
    csv.register_dialect('PARCEL_dialect',
                         delimiter=',',
                         quoting=csv.QUOTE_ALL,
                         skipinitialspace=True
                         )
    with open(dir_file, mode='r') as csv_file:
        csv_reader = csv.DictReader(csv_file, dialect='PARCEL_dialect')
        for row in csv_reader:
            r = dict()  # r aka record
            for key, value in column_mapping.items():
                r[value] = row.get(key)
            raw_v_rows.append(r)
    return raw_v_rows


def load_abi(file_name):
    fp = f'{HOME_PACKAGE}/{file_name}'
    with open(fp, 'r') as f:
        abi = json.load(f)
        return abi


def force2bool(input : str or bool) -> bool:
    if isinstance(input, bool):
        return input

    elif isinstance(input, str):

        if input.lower().strip() == 'true': return True
        if input.lower().strip() == 't':    return True
        if input.lower().strip() == 'True': return True
        if input.lower().strip() == 'T':    return True
        if input.lower().strip() == 'yes':  return True
        if input.lower().strip() == 'Yes':  return True
        if input.lower().strip() == 'y':    return True
        return False

    else:
        return False


def force_int(value, default=0):
    try:
        return int(value)
    except Exception as _e:
        return default


def force_float(value, default=0.0):
    try:
        return float(value)
    except Exception as _e:
        return default


def randomword(length):
   letters = string.ascii_lowercase
   return ''.join(random.choice(letters) for i in range(length))


def file_latest_in_path(log_dir: str = None) -> str:
    if log_dir is None:
        log_dir = HOME_TMP

    files = os.listdir(log_dir)
    files.sort(key=lambda x: os.path.getmtime(os.path.join(log_dir, x)))
    if len(files) == 0:
        return None
    return os.path.join(log_dir, files[-1])


def find_latest_row_index_log(file_report) -> tuple:
    df = pd.read_csv(file_report)
    # index last row
    index = df.index[-1]
    row = df.loc[index]
    return index, row


def df_to_csv(df, file_path):
    # Save dataframe as csv.
    df.to_csv(file_path, index=False)


def csv_to_df(file_path):
    # Read csv to dataframe.
    df = pd.read_csv(file_path)
    return df


def add_to_csv(file_path, add_text):
    # Add a line to file_name.csv
    # Should be like [xx,xx,xx]
    df = csv_to_df(file_path)
    l = len(df)
    df.loc[l] = add_text
    df_to_csv(df, file_path)


def get_fa_backup(str_fab) -> str:
    co = str_fab.replace('\n', ' ').split(' ')
    fab = ''
    for text in co:
        if len(text) == 12:
            fab = text
            break
    return fab


def ip():
    try:
        ip = requests.get('https://checkip.amazonaws.com').text.strip()
    except Exception as _e:
        ip = ""
    return ip


def get_ip():
    ip_address_now = ip()
    # check file exist
    file_path = f"{HOME_TMP}/../ip_address.txt"
    if not os.path.exists(file_path):
        with open(file_path, "w") as f:
            f.write(f"{ip_address_now}|{ip_address_now}")
    else:
        # get last ip from file
        with open(file_path, "r") as f:
            line = f.read().replace("\n", "")
            current_ip = line.split("|")[1]

        # compare
        if current_ip != ip_address_now:
            logger.info(f"Last IP Address: {current_ip}")
            # write to file
            with open(file_path, "w") as f:
                f.write(f"{current_ip}|{ip_address_now}")

    logger.info(f"IP Address: {ip_address_now}")
    return ip_address_now


def cook_address(address: str) -> str:
    """ Return cooked address """
    address = address.lower().strip().replace("0x", "").replace("0:", "")
    return address


def user_data_dir(address: str = None) -> str or None:
    """ Return user data dir """
    if address is None and USER_DATA_DIR is None:
        return None

    address = cook_address(address)  # cook address
    # specific user data dir if set
    # create user data dir if not exist by using address of wallet
    udd = USER_DATA_DIR if USER_DATA_DIR else os.path.join(ALL_USER_DATA_DIR, address)

    if not os.path.exists(udd):
        os.makedirs(udd)

    return udd


def totp(secret: str) -> str:
    """ Calculate TOTP using time and key """
    key = base64.b32decode(secret, True)
    now = int(time.time() // 30)
    msg = now.to_bytes(8, "big")
    digest = hmac.new(key, msg, "sha1").digest()
    offset = digest[19] & 0xF
    code = digest[offset : offset + 4]
    code = int.from_bytes(code, "big") & 0x7FFFFFFF
    code = code % 1000000
    return "{:06d}".format(code)


def reboot():
    """ Reboot dongle """
    Dongle().reboot()


def reboot_reconnect():
    """ Reboot dongle """
    logger.info("Reboot dongle")
    current_network = get_ssid()
    reboot()
    time.sleep(50)
    res = None
    while not res:
        try:
            res = ControlConnection(wifi_ssid=current_network, wifi_password=NETWORK_PASSWORD).wifi_connector()
        except Exception as _e:
            logger.error(f"Error connect {current_network}: {_e} retry after 10s")
        time.sleep(10)
    time.sleep(20)
    logger.info(f"New IP Address: {ip()}")


def change_network():
    """ Change network """
    try:
        logger.info(f"IP Address:     {ip()}")
        change_to_network = None
        while not change_to_network:
            try:
                change_to_network = get_network()
            except Exception as _e:
                logger.error(f"Error get network: {_e}, retry after 3s")
            time.sleep(3)
        logger.info(f"Change from {get_ssid()} to {change_to_network}")

        reboot()

        res = None
        while not res:
            try:
                res = ControlConnection(wifi_ssid=change_to_network, wifi_password=NETWORK_PASSWORD).wifi_connector()
            except Exception as _e:
                logger.error(f"Error connect {change_to_network}: {_e} retry after 10s")
            time.sleep(10)

        logger.info(f"New IP Address: {ip()}")
    except Exception as e:
        logger.error(f"Error change network: {e}")


def get_ssid():
    """Get the SSID of the connected WiFi."""
    process = subprocess.Popen([PATH_OF_AIRPORT, "-I"], stdout=subprocess.PIPE)
    out, err = process.communicate()
    process.wait()
    output = {}
    for line in out.decode("utf-8").split("\n"):
        if ": " in line:
            key, value = line.split(": ")
            key = key.strip()
            value = value.strip()
            output[key] = value

    return output["SSID"]


def get_network(exclude_network: str = None) -> str:
    """ Get network """
    if exclude_network is None:
        exclude_network = get_ssid()

    list_network = copy.deepcopy(LIST_NETWORK)
    if exclude_network in list_network:
        list_network.remove(exclude_network)
    logger.info(f"List network: {list_network}")
    network = list_network[0]
    return network
